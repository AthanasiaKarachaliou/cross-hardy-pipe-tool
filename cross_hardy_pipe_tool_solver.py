from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import math
import os
import tempfile

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


@dataclass
class Node:
    node_id: str
    demand: float = 0.0
    is_source: bool = False
    source_pressure: Optional[float] = None
    x: Optional[float] = None
    y: Optional[float] = None
    pressure: Optional[float] = None


@dataclass
class Pipe:
    pipe_id: str
    from_node: str
    to_node: str
    length: float
    diameter: float
    K: float
    loss_exponent: float = 2.0
    status: str = 'OPEN'
    flow: float = 0.0
    dP_signed: float = 0.0
    dP_abs: float = 0.0
    velocity: float = 0.0
    loops: str = ''


def _truthy(value) -> bool:
    return str(value).strip().upper() in {'TRUE', '1', 'YES', 'Y', 'OPEN'}


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _read_sheet_autodetect(
    xlsx_path: str | Path,
    sheet_name: str,
    required_columns: List[str],
    search_rows: int = 10,
) -> pd.DataFrame:
    for header_row in range(search_rows):
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=header_row)
            df = _clean_columns(df)
            if all(col in df.columns for col in required_columns):
                return df
        except Exception:
            continue
    raise ValueError(
        f"Could not find required columns {required_columns} in sheet '{sheet_name}'. Check the sheet name and header row."
    )


def _pick_first_existing(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for col in candidates:
        if col in df.columns:
            return col
    return None


class PipeNetwork:
    def __init__(self, nodes: Dict[str, Node], pipes: Dict[str, Pipe], metadata: Optional[dict] = None):
        self.nodes = nodes
        self.pipes = pipes
        self.metadata = metadata or {}
        self.adj = self._build_adjacency()

    def _build_adjacency(self) -> Dict[str, List[Tuple[str, str]]]:
        adj = {nid: [] for nid in self.nodes}
        for pipe in self.pipes.values():
            if str(pipe.status).upper() != 'OPEN':
                continue
            if pipe.from_node in self.nodes and pipe.to_node in self.nodes:
                adj[pipe.from_node].append((pipe.to_node, pipe.pipe_id))
                adj[pipe.to_node].append((pipe.from_node, pipe.pipe_id))
        return adj

    def _dfs_nodes(self, start: str) -> Set[str]:
        stack = [start]
        visited: Set[str] = set()
        while stack:
            node = stack.pop()
            if node in visited:
                continue
            visited.add(node)
            for nbr, _ in self.adj.get(node, []):
                if nbr not in visited:
                    stack.append(nbr)
        return visited

    def validate(self) -> None:
        if not self.nodes:
            raise ValueError('No nodes were loaded.')
        if not self.pipes:
            raise ValueError('No pipes were loaded.')

        open_pipes = [p for p in self.pipes.values() if str(p.status).upper() == 'OPEN']
        if not open_pipes:
            raise ValueError('No OPEN pipes were found.')

        for pipe in open_pipes:
            if pipe.from_node not in self.nodes or pipe.to_node not in self.nodes:
                raise ValueError(f'Pipe {pipe.pipe_id} references a missing node.')

        sources = [n for n in self.nodes.values() if n.is_source]
        if not sources:
            raise ValueError('At least one source node is required.')

        ref_sources = [n for n in sources if n.source_pressure is not None]
        if not ref_sources:
            raise ValueError('At least one source node must have a source_pressure.')

        start = ref_sources[0].node_id
        visited = self._dfs_nodes(start)
        if len(visited) != len(self.nodes):
            missing = sorted(set(self.nodes) - visited)
            raise ValueError(f'The active network is disconnected. Unreached nodes: {missing}')

    def get_reference_source(self) -> Node:
        candidates = [n for n in self.nodes.values() if n.is_source and n.source_pressure is not None]
        if not candidates:
            raise ValueError('No source node with source_pressure found.')
        return candidates[0]

    def build_bfs_tree(self, root: str) -> Tuple[Dict[str, Optional[str]], List[str]]:
        parent: Dict[str, Optional[str]] = {root: None}
        order = [root]
        queue = [root]
        while queue:
            u = queue.pop(0)
            for v, _ in self.adj.get(u, []):
                if v not in parent:
                    parent[v] = u
                    queue.append(v)
                    order.append(v)
        return parent, order

    def edge_pipe_id(self, u: str, v: str) -> str:
        for nbr, pid in self.adj.get(u, []):
            if nbr == v:
                return pid
        raise KeyError(f'No active pipe found between {u} and {v}')

    def assign_initial_tree_flows(self) -> None:
        src = self.get_reference_source().node_id
        parent, order = self.build_bfs_tree(src)
        subtree_demand = {nid: self.nodes[nid].demand for nid in self.nodes}

        for nid in reversed(order):
            p = parent[nid]
            if p is not None:
                subtree_demand[p] += subtree_demand[nid]

        tree_pipe_ids: Set[str] = set()
        for nid in order:
            p = parent[nid]
            if p is None:
                continue
            pid = self.edge_pipe_id(p, nid)
            tree_pipe_ids.add(pid)
            pipe = self.pipes[pid]
            q = subtree_demand[nid]
            if pipe.from_node == p and pipe.to_node == nid:
                pipe.flow = q
            elif pipe.from_node == nid and pipe.to_node == p:
                pipe.flow = -q
            else:
                raise RuntimeError(f'Tree edge ({p}, {nid}) does not align with pipe {pid}')

        for pid, pipe in self.pipes.items():
            if str(pipe.status).upper() == 'OPEN' and pid not in tree_pipe_ids:
                pipe.flow = 0.0

    @staticmethod
    def pressure_loss_signed(pipe: Pipe, q: float) -> float:
        if q == 0.0:
            return 0.0
        n = pipe.loss_exponent
        return pipe.K * q * (abs(q) ** (n - 1.0))

    @staticmethod
    def velocity(pipe: Pipe, q: float) -> float:
        area = math.pi * (pipe.diameter ** 2) / 4.0
        if area <= 0:
            return 0.0
        return abs(q) / area

    def _path_to_root(self, node: str, parent: Dict[str, Optional[str]]) -> List[str]:
        path = []
        cur = node
        while cur is not None:
            path.append(cur)
            cur = parent[cur]
        return path

    def find_fundamental_loops(self) -> List[List[Tuple[str, str, str]]]:
        root = self.get_reference_source().node_id
        parent, _ = self.build_bfs_tree(root)

        tree_edges = set()
        for child, p in parent.items():
            if p is not None:
                tree_edges.add(frozenset((child, p)))

        seen_edges = set()
        non_tree_edges = []
        for pid, pipe in self.pipes.items():
            if str(pipe.status).upper() != 'OPEN':
                continue
            edge_key = frozenset((pipe.from_node, pipe.to_node))
            if edge_key in seen_edges:
                continue
            seen_edges.add(edge_key)
            if edge_key not in tree_edges:
                non_tree_edges.append((pipe.from_node, pipe.to_node, pid))

        loops: List[List[Tuple[str, str, str]]] = []
        for u, v, _ in non_tree_edges:
            path_u = self._path_to_root(u, parent)
            path_v = self._path_to_root(v, parent)
            set_v = set(path_v)
            lca = next(node for node in path_u if node in set_v)

            up_u = []
            cur = u
            while cur != lca:
                up_u.append(cur)
                cur = parent[cur]
            up_u.append(lca)

            down_v = []
            cur = v
            while cur != lca:
                down_v.append(cur)
                cur = parent[cur]
            down_v.append(lca)
            down_v.reverse()

            node_cycle = up_u + down_v[1:] + [u]
            if len(node_cycle) < 4:
                continue

            loop_edges: List[Tuple[str, str, str]] = []
            for a, b in zip(node_cycle[:-1], node_cycle[1:]):
                pid = self.edge_pipe_id(a, b)
                loop_edges.append((pid, a, b))
            loops.append(loop_edges)

        for pipe in self.pipes.values():
            pipe.loops = ''
        for i, loop in enumerate(loops, start=1):
            members = sorted({pid for pid, _, _ in loop})
            for pid in members:
                pipe = self.pipes[pid]
                pipe.loops = (pipe.loops + ',' if pipe.loops else '') + f'L{i}'

        return loops

    def _loop_correction(self, loop_edges: List[Tuple[str, str, str]]) -> float:
        num = 0.0
        den = 0.0
        for pid, lu, lv in loop_edges:
            pipe = self.pipes[pid]
            q = pipe.flow if abs(pipe.flow) > 1e-12 else 1e-12
            if pipe.from_node == lu and pipe.to_node == lv:
                s = 1.0
            elif pipe.from_node == lv and pipe.to_node == lu:
                s = -1.0
            else:
                raise RuntimeError(f'Loop edge mismatch for pipe {pid}')
            n = pipe.loss_exponent
            k = pipe.K
            num += s * k * q * (abs(q) ** (n - 1.0))
            den += n * k * (abs(q) ** (n - 1.0))
        return 0.0 if abs(den) < 1e-18 else -num / den

    def run_hardy_cross(self, max_iters: int = 30, tol_flow_change: float = 1e-6) -> dict:
        loops = self.find_fundamental_loops()
        if not loops:
            return {'iterations': 0, 'converged': True, 'max_delta_q': 0.0, 'loop_count': 0}

        converged = False
        final_max = 0.0
        it = 0
        for it in range(1, max_iters + 1):
            max_dq = 0.0
            for loop_edges in loops:
                dq = self._loop_correction(loop_edges)
                for pid, lu, lv in loop_edges:
                    pipe = self.pipes[pid]
                    sign = 1.0 if (pipe.from_node == lu and pipe.to_node == lv) else -1.0
                    pipe.flow += sign * dq
                max_dq = max(max_dq, abs(dq))
            final_max = max_dq
            if max_dq < tol_flow_change:
                converged = True
                break

        return {
            'iterations': it,
            'converged': converged,
            'max_delta_q': final_max,
            'loop_count': len(loops),
        }

    def postprocess(self) -> dict:
        for pipe in self.pipes.values():
            pipe.dP_signed = self.pressure_loss_signed(pipe, pipe.flow)
            pipe.dP_abs = abs(pipe.dP_signed)
            pipe.velocity = self.velocity(pipe, pipe.flow)

        ref = self.get_reference_source()
        for node in self.nodes.values():
            node.pressure = None
        ref.pressure = ref.source_pressure

        parent, order = self.build_bfs_tree(ref.node_id)
        for nid in order:
            p = parent[nid]
            if p is None:
                continue
            pid = self.edge_pipe_id(p, nid)
            pipe = self.pipes[pid]
            parent_pressure = self.nodes[p].pressure
            if parent_pressure is None:
                raise RuntimeError(f'Pressure for parent node {p} is not known.')

            if pipe.flow >= 0:
                upstream = pipe.from_node
                downstream = pipe.to_node
            else:
                upstream = pipe.to_node
                downstream = pipe.from_node

            if p == upstream:
                self.nodes[nid].pressure = parent_pressure - pipe.dP_abs
            elif p == downstream:
                self.nodes[nid].pressure = parent_pressure + pipe.dP_abs
            else:
                raise RuntimeError(f'Tree edge ({p}, {nid}) does not align with pipe {pid}')

        drops = {
            nid: ref.source_pressure - node.pressure
            for nid, node in self.nodes.items()
            if node.pressure is not None
        }

        worst_pressure_node = min(
            [n for n in self.nodes.values() if n.pressure is not None],
            key=lambda n: n.pressure,
        )
        worst_drop_node_id, worst_drop = max(drops.items(), key=lambda x: x[1])
        max_vel_pipe = max(self.pipes.values(), key=lambda p: p.velocity)

        return {
            'reference_source': ref.node_id,
            'reference_pressure': ref.source_pressure,
            'worst_node_pressure_id': worst_pressure_node.node_id,
            'worst_node_pressure': worst_pressure_node.pressure,
            'worst_node_drop_id': worst_drop_node_id,
            'worst_node_drop': worst_drop,
            'max_velocity_pipe_id': max_vel_pipe.pipe_id,
            'max_velocity': max_vel_pipe.velocity,
        }

    def nodes_dataframe(self) -> pd.DataFrame:
        ref = self.get_reference_source()
        rows = []
        for node in self.nodes.values():
            rows.append({
                'node_id': node.node_id,
                'demand': node.demand,
                'is_source': node.is_source,
                'source_pressure': node.source_pressure,
                'pressure': node.pressure,
                'drop_from_reference': (ref.source_pressure - node.pressure) if node.pressure is not None else None,
                'x': node.x,
                'y': node.y,
            })
        return pd.DataFrame(rows)

    def pipes_dataframe(self) -> pd.DataFrame:
        rows = []
        for pipe in self.pipes.values():
            rows.append({
                'pipe_id': pipe.pipe_id,
                'from_node': pipe.from_node,
                'to_node': pipe.to_node,
                'length': pipe.length,
                'diameter': pipe.diameter,
                'K': pipe.K,
                'loss_exponent': pipe.loss_exponent,
                'status': pipe.status,
                'flow': pipe.flow,
                'dP_signed': pipe.dP_signed,
                'dP_abs': pipe.dP_abs,
                'velocity': pipe.velocity,
                'direction': f"{pipe.from_node}->{pipe.to_node}" if pipe.flow >= 0 else f"{pipe.to_node}->{pipe.from_node}",
                'loops': pipe.loops,
            })
        return pd.DataFrame(rows)

    def create_map_bytes(self) -> bytes:
        positions = {}
        have_xy = all(self.nodes[n].x is not None and self.nodes[n].y is not None for n in self.nodes)
        if have_xy:
            for nid, node in self.nodes.items():
                positions[nid] = (node.x, node.y)
        else:
            n = len(self.nodes)
            radius = max(5.0, n)
            sorted_nodes = sorted(self.nodes)
            for i, nid in enumerate(sorted_nodes):
                theta = 2.0 * math.pi * i / max(1, n)
                positions[nid] = (radius * math.cos(theta), radius * math.sin(theta))

        fig = plt.figure(figsize=(12, 8))
        xs = [positions[n][0] for n in self.nodes]
        ys = [positions[n][1] for n in self.nodes]
        plt.scatter(xs, ys, s=500, c='#d7ebff', edgecolors='#1f4e79', zorder=3)

        for nid, (x, y) in positions.items():
            plt.text(x, y, nid, ha='center', va='center', fontsize=8, zorder=4)

        vel_values = [self.pipes[pid].velocity for pid in self.pipes] or [0.0]
        vmin = min(vel_values)
        vmax = max(vel_values)
        span = max(vmax - vmin, 1e-9)

        for pid, pipe in self.pipes.items():
            if pipe.from_node not in positions or pipe.to_node not in positions:
                continue
            x1, y1 = positions[pipe.from_node]
            x2, y2 = positions[pipe.to_node]
            t = (pipe.velocity - vmin) / span
            color = plt.cm.plasma(t)
            lw = 1.5 + min(4.0, abs(pipe.flow) / 10.0)
            plt.plot([x1, x2], [y1, y2], color=color, linewidth=lw, zorder=1)
            mx, my = (x1 + x2) / 2.0, (y1 + y2) / 2.0
            label = f"{pid}\nQ={pipe.flow:.2f}\nv={pipe.velocity:.2f}"
            plt.text(mx, my, label, fontsize=6, bbox=dict(boxstyle='round,pad=0.15', fc='white', ec='none', alpha=0.8), zorder=5)

            dx, dy = x2 - x1, y2 - y1
            if pipe.flow < 0:
                dx, dy = -dx, -dy
                x1, y1 = x2, y2
            arrow_len = 0.35
            norm = math.hypot(dx, dy)
            if norm > 0:
                ux, uy = dx / norm, dy / norm
                ax = x1 + 0.55 * dx
                ay = y1 + 0.55 * dy
                plt.arrow(ax, ay, arrow_len * ux, arrow_len * uy, head_width=0.25, head_length=0.25, fc=color, ec=color, length_includes_head=True, zorder=2)

        plt.title('Pipe Network Map')
        plt.axis('equal')
        plt.axis('off')
        plt.tight_layout()

        buffer = BytesIO()
        fig.savefig(buffer, format='png', dpi=200)
        plt.close(fig)
        buffer.seek(0)
        return buffer.getvalue()


def load_network_from_workbook(xlsx_path: str | Path) -> PipeNetwork:
    xlsx_path = str(xlsx_path)
    nodes_df = _read_sheet_autodetect(xlsx_path, 'Nodes', ['Node_ID'])
    seg_df = _read_sheet_autodetect(xlsx_path, 'Segments', ['Segment_ID', 'From_Node', 'To_Node'])
    loads_df = _read_sheet_autodetect(xlsx_path, 'Loads', ['Node_ID'])
    sources_df = _read_sheet_autodetect(xlsx_path, 'Sources', ['Node_ID'])

    node_active_col = _pick_first_existing(nodes_df, ['Active', 'ACTIVE', 'Status'])
    active_nodes_df = nodes_df[nodes_df[node_active_col].apply(_truthy)].copy() if node_active_col else nodes_df.copy()

    source_active_col = _pick_first_existing(sources_df, ['Active', 'ACTIVE', 'Status'])
    active_sources = sources_df[sources_df[source_active_col].apply(_truthy)].copy() if source_active_col else sources_df.copy()

    source_pressure_col = _pick_first_existing(sources_df, ['Pressure_psig', 'source_pressure', 'Source_Pressure', 'Pressure'])
    if source_pressure_col is None:
        raise ValueError('Sources sheet must contain a pressure column such as Pressure_psig.')
    source_pressures = {
        str(r['Node_ID']): float(r[source_pressure_col])
        for _, r in active_sources.iterrows()
        if pd.notna(r[source_pressure_col])
    }

    load_active_col = _pick_first_existing(loads_df, ['Active', 'ACTIVE', 'Status'])
    active_loads = loads_df[loads_df[load_active_col].apply(_truthy)].copy() if load_active_col else loads_df.copy()
    demand_col = _pick_first_existing(active_loads, ['Design_Demand_scfm', 'Peak_Demand_scfm', 'Demand', 'demand'])
    if demand_col is None:
        active_loads['Demand'] = 0.0
        demand_col = 'Demand'
    load_by_node = active_loads.groupby('Node_ID')[demand_col].sum().to_dict() if len(active_loads) else {}

    x_col = _pick_first_existing(active_nodes_df, ['X', 'x'])
    y_col = _pick_first_existing(active_nodes_df, ['Y', 'y'])
    nodes: Dict[str, Node] = {}
    for _, r in active_nodes_df.iterrows():
        nid = str(r['Node_ID'])
        nodes[nid] = Node(
            node_id=nid,
            demand=float(load_by_node.get(nid, 0.0)),
            is_source=nid in source_pressures,
            source_pressure=source_pressures.get(nid),
            x=float(r[x_col]) if x_col and pd.notna(r.get(x_col, None)) else None,
            y=float(r[y_col]) if y_col and pd.notna(r.get(y_col, None)) else None,
        )

    seg_active_col = _pick_first_existing(seg_df, ['Active', 'ACTIVE', 'Status'])
    active_segments = seg_df[seg_df[seg_active_col].apply(_truthy)].copy() if seg_active_col else seg_df.copy()

    k_col = _pick_first_existing(active_segments, ['K_psi_per_scfm2_gas_corrected', 'K', 'Resistance_K', 'k'])
    if k_col is None:
        raise ValueError('Segments sheet must contain a K column such as K_psi_per_scfm2_gas_corrected or K.')

    length_col = _pick_first_existing(active_segments, ['Eq_Length_ft', 'Length_ft', 'length', 'Length'])
    diameter_col = _pick_first_existing(active_segments, ['ID_in', 'Size_in', 'diameter', 'Diameter'])
    exponent_col = _pick_first_existing(active_segments, ['loss_exponent', 'Loss_Exponent', 'n'])

    pipes: Dict[str, Pipe] = {}
    for _, r in active_segments.iterrows():
        pid = str(r['Segment_ID'])
        if pd.isna(r[k_col]):
            raise ValueError(f'Segment {pid} is missing a K value.')
        pipes[pid] = Pipe(
            pipe_id=pid,
            from_node=str(r['From_Node']),
            to_node=str(r['To_Node']),
            length=float(r[length_col]) if length_col and pd.notna(r.get(length_col, None)) else 0.0,
            diameter=float(r[diameter_col]) if diameter_col and pd.notna(r.get(diameter_col, None)) else 0.0,
            K=float(r[k_col]),
            loss_exponent=float(r[exponent_col]) if exponent_col and pd.notna(r.get(exponent_col, None)) else 2.0,
            status='OPEN',
        )

    return PipeNetwork(nodes, pipes, metadata={'input_mode': 'xlsx', 'xlsx': xlsx_path})


def load_review_tables_from_workbook(xlsx_path: str | Path) -> dict:
    xlsx_path = str(xlsx_path)
    nodes_df = _read_sheet_autodetect(xlsx_path, 'Nodes', ['Node_ID'])
    segments_df = _read_sheet_autodetect(xlsx_path, 'Segments', ['Segment_ID', 'From_Node', 'To_Node'])
    loads_df = _read_sheet_autodetect(xlsx_path, 'Loads', ['Node_ID'])
    sources_df = _read_sheet_autodetect(xlsx_path, 'Sources', ['Node_ID'])

    return {
        'nodes': nodes_df.copy(),
        'segments': segments_df.copy(),
        'loads': loads_df.copy(),
        'sources': sources_df.copy(),
    }


def _active_mask(df: pd.DataFrame, active_col: str = 'Active') -> pd.Series:
    if active_col not in df.columns:
        return pd.Series([True] * len(df), index=df.index)
    return df[active_col].apply(_truthy)


def _required_columns_present(df: pd.DataFrame, required: list[str]) -> list[str]:
    return [col for col in required if col not in df.columns]


def validate_draft_tables(nodes_df: pd.DataFrame, segments_df: pd.DataFrame, loads_df: pd.DataFrame, sources_df: pd.DataFrame) -> dict:
    errors = []
    warnings = []

    node_required = ['Node_ID', 'Active']
    segment_required = ['Segment_ID', 'From_Node', 'To_Node', 'Active', 'K_psi_per_scfm2_gas_corrected']
    load_required = ['Load_ID', 'Node_ID', 'Active']
    source_required = ['Source_ID', 'Node_ID', 'Pressure_psig', 'Active']

    missing = _required_columns_present(nodes_df, node_required)
    if missing:
        errors.append(f'Nodes sheet is missing columns: {missing}')
    missing = _required_columns_present(segments_df, segment_required)
    if missing:
        errors.append(f'Segments sheet is missing columns: {missing}')
    missing = _required_columns_present(loads_df, load_required)
    if missing:
        errors.append(f'Loads sheet is missing columns: {missing}')
    missing = _required_columns_present(sources_df, source_required)
    if missing:
        errors.append(f'Sources sheet is missing columns: {missing}')

    if errors:
        return {'errors': errors, 'warnings': warnings, 'stats': {}}

    active_nodes = nodes_df[_active_mask(nodes_df)].copy()
    active_segments = segments_df[_active_mask(segments_df)].copy()
    active_loads = loads_df[_active_mask(loads_df)].copy()
    active_sources = sources_df[_active_mask(sources_df)].copy()

    if active_nodes.empty:
        errors.append('There are no active nodes.')
    if active_segments.empty:
        errors.append('There are no active segments.')
    if active_sources.empty:
        errors.append('There are no active sources.')

    if active_nodes['Node_ID'].duplicated().any():
        dupes = active_nodes.loc[active_nodes['Node_ID'].duplicated(), 'Node_ID'].tolist()
        errors.append(f'Duplicate active Node_ID values found: {dupes}')

    active_node_ids = set(active_nodes['Node_ID'].astype(str))

    if not active_segments[~active_segments['From_Node'].astype(str).isin(active_node_ids)].empty:
        errors.append('Some active segments have From_Node values that do not exist in active nodes.')
    if not active_segments[~active_segments['To_Node'].astype(str).isin(active_node_ids)].empty:
        errors.append('Some active segments have To_Node values that do not exist in active nodes.')
    if not active_sources[~active_sources['Node_ID'].astype(str).isin(active_node_ids)].empty:
        errors.append('Some active sources reference Node_ID values not found in active nodes.')
    if not active_loads[~active_loads['Node_ID'].astype(str).isin(active_node_ids)].empty:
        errors.append('Some active loads reference Node_ID values not found in active nodes.')

    missing_k = active_segments[
        active_segments['K_psi_per_scfm2_gas_corrected'].isna() |
        (active_segments['K_psi_per_scfm2_gas_corrected'].astype(str).str.strip() == '')
    ]
    if not missing_k.empty:
        errors.append(f"{len(missing_k)} active segment(s) are missing K_psi_per_scfm2_gas_corrected.")

    missing_source_pressure = active_sources[
        active_sources['Pressure_psig'].isna() |
        (active_sources['Pressure_psig'].astype(str).str.strip() == '')
    ]
    if not missing_source_pressure.empty:
        errors.append('At least one active source is missing Pressure_psig.')

    if 'Design_Demand_scfm' in active_loads.columns:
        empty_design = active_loads[
            active_loads['Design_Demand_scfm'].isna() |
            (active_loads['Design_Demand_scfm'].astype(str).str.strip() == '')
        ]
        if not empty_design.empty:
            warnings.append(f"{len(empty_design)} active load(s) do not have Design_Demand_scfm yet.")

    try:
        source_pressures = {
            str(r['Node_ID']): float(r['Pressure_psig'])
            for _, r in active_sources.iterrows()
            if pd.notna(r['Pressure_psig']) and str(r['Pressure_psig']).strip() != ''
        }

        load_by_node = {}
        if 'Design_Demand_scfm' in active_loads.columns:
            tmp = active_loads.copy()
            tmp['Design_Demand_scfm'] = pd.to_numeric(tmp['Design_Demand_scfm'], errors='coerce').fillna(0.0)
            load_by_node = tmp.groupby('Node_ID')['Design_Demand_scfm'].sum().to_dict()

        x_col = 'X' if 'X' in active_nodes.columns else None
        y_col = 'Y' if 'Y' in active_nodes.columns else None

        temp_nodes = {}
        for _, r in active_nodes.iterrows():
            nid = str(r['Node_ID'])
            temp_nodes[nid] = Node(
                node_id=nid,
                demand=float(load_by_node.get(nid, 0.0)),
                is_source=nid in source_pressures,
                source_pressure=source_pressures.get(nid),
                x=float(r[x_col]) if x_col and pd.notna(r.get(x_col)) else None,
                y=float(r[y_col]) if y_col and pd.notna(r.get(y_col)) else None,
            )

        temp_pipes = {}
        for _, r in active_segments.iterrows():
            pid = str(r['Segment_ID'])
            k_val = r['K_psi_per_scfm2_gas_corrected']
            k_float = 1.0 if pd.isna(k_val) or str(k_val).strip() == '' else float(k_val)

            length_val = 0.0
            if 'Eq_Length_ft' in active_segments.columns and pd.notna(r.get('Eq_Length_ft')):
                length_val = float(r['Eq_Length_ft'])
            elif 'Length_ft' in active_segments.columns and pd.notna(r.get('Length_ft')):
                length_val = float(r['Length_ft'])

            diameter_val = 0.0
            if 'ID_in' in active_segments.columns and pd.notna(r.get('ID_in')):
                diameter_val = float(r['ID_in'])
            elif 'Size_in' in active_segments.columns and pd.notna(r.get('Size_in')):
                diameter_val = float(r['Size_in'])

            temp_pipes[pid] = Pipe(
                pipe_id=pid,
                from_node=str(r['From_Node']),
                to_node=str(r['To_Node']),
                length=length_val,
                diameter=diameter_val,
                K=k_float,
                loss_exponent=2.0,
                status='OPEN',
            )

        PipeNetwork(temp_nodes, temp_pipes).validate()
    except Exception as exc:
        errors.append(f'Network validation failed: {exc}')

    stats = {
        'active_nodes': len(active_nodes),
        'active_segments': len(active_segments),
        'active_loads': len(active_loads),
        'active_sources': len(active_sources),
    }
    return {'errors': errors, 'warnings': warnings, 'stats': stats}


def _clear_sheet_rows(ws, start_row: int = 2) -> None:
    max_rows = ws.max_row
    if max_rows >= start_row:
        ws.delete_rows(start_row, max_rows - start_row + 1)


def _sanitize_for_excel(df: pd.DataFrame, allowed_columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    keep = [c for c in allowed_columns if c in out.columns]
    return out[keep].copy()


def export_reviewed_draft_workbook(template_path: str | Path, nodes_df: pd.DataFrame, segments_df: pd.DataFrame, loads_df: pd.DataFrame, sources_df: pd.DataFrame) -> bytes:
    wb = load_workbook(template_path)

    nodes_cols = ['Node_ID', 'Description', 'Node_Type', 'X', 'Y', 'Active']
    segments_cols = [
        'Segment_ID', 'Type', 'From_Node', 'To_Node', 'Description', 'Size_in',
        'Length_ft', 'Fittings_ft', 'Eq_Length_ft', 'Active', 'ID_in',
        'K_psi_per_scfm2_gas_corrected', 'Max_Vel_Limit_ft_s'
    ]
    loads_cols = ['Load_ID', 'Equipment', 'Node_ID', 'Peak_Demand_scfm', 'Active', 'Design_Demand_scfm']
    sources_cols = ['Source_ID', 'Node_ID', 'Pressure_psig', 'Active', 'Description']

    data_map = {
        'Nodes': _sanitize_for_excel(nodes_df, nodes_cols),
        'Segments': _sanitize_for_excel(segments_df, segments_cols),
        'Loads': _sanitize_for_excel(loads_df, loads_cols),
        'Sources': _sanitize_for_excel(sources_df, sources_cols),
    }

    for sheet_name, df in data_map.items():
        ws = wb[sheet_name]
        _clear_sheet_rows(ws, start_row=2)
        for _, row in df.iterrows():
            values = [None if pd.isna(v) else v for v in row.tolist()]
            ws.append(values)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def solve_workbook(xlsx_file_or_path: str | Path, max_iters: int = 30, tol: float = 1e-6):
    net = load_network_from_workbook(xlsx_file_or_path)
    net.validate()
    net.assign_initial_tree_flows()
    run_info = net.run_hardy_cross(max_iters=max_iters, tol_flow_change=tol)
    summary = net.postprocess()
    summary['loop_count'] = run_info['loop_count']
    summary['iterations'] = run_info['iterations']
    summary['converged'] = run_info['converged']
    summary['final_max_dq'] = run_info['max_delta_q']
    nodes_df = net.nodes_dataframe()
    pipes_df = net.pipes_dataframe()
    map_bytes = net.create_map_bytes()
    return summary, nodes_df, pipes_df, map_bytes


def solve_reviewed_tables(template_path: str | Path, nodes_df: pd.DataFrame, segments_df: pd.DataFrame, loads_df: pd.DataFrame, sources_df: pd.DataFrame, max_iters: int = 30, tol: float = 1e-6):
    workbook_bytes = export_reviewed_draft_workbook(
        template_path=template_path,
        nodes_df=nodes_df,
        segments_df=segments_df,
        loads_df=loads_df,
        sources_df=sources_df,
    )

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(workbook_bytes)
        temp_path = tmp.name

    try:
        return solve_workbook(temp_path, max_iters=max_iters, tol=tol), workbook_bytes
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
